# -*- encoding: utf-8 -*-

"""


"""
import time

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch, cm, mm
from reportlab.lib import colors

from reportlab.graphics.barcode.code39 import Standard39
from reportlab.graphics.barcode.code128 import Code128
from reportlab.graphics.barcode import qr

from reportlab.graphics.shapes import Drawing 
from reportlab.pdfgen.canvas import Canvas

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont 

from openpyxl import load_workbook

pdfmetrics.registerFont(TTFont('simsun', 'simsun.ttc'))

from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate, NextPageTemplate
from reportlab.platypus import Paragraph, Spacer, ParaLines, PageBreak, Table, TableStyle

def get_data(wb, i):
        
    s = 'Sheet{}'.format(i)
    sheet = wb[s]
    
    rows = []
    d = {}
    
    O = sheet.cell(2, 1).value

    V = sheet.cell(2, 4).value
    
    for row in range(2,sheet.max_row+1):    

        if sheet.cell(row, 4).value != V or sheet.cell(row, 1).value == None:
            key = "{}-{}".format(O, V)
            d[key]= rows
            rows = []
            V = sheet.cell(row, 4).value
        if sheet.cell(row, 1).value == None:
            break
        #steps = {}
        cell_vals = []
        for col in range(1, 11):
            cell = sheet.cell(row=row, column=col)
            if type(cell.value) == str:
                val= cell.value#.encode("utf-8")
            else:
                val=cell.value
            cell_vals.append(val)
        rows.append(cell_vals)
    return d
    
    
def addPageNumber(canvas, doc):
    """
    Add the page number
    """
    page_num = canvas.getPageNumber()
    text = "Page #%s" % page_num
    canvas.drawRightString(200*mm, 20*mm, text)


    
class PartListReport(BaseDocTemplate):
    
    def __init__(self, step, vals, **kwargs):
        
        filename = "pdfs\\{}.pdf".format(step)
        self.filename = filename
        self.vals = vals
        super().__init__(filename, page_size=A4, _pageBreakQuick=0, **kwargs)
        
        self.run()
        
        
    def run(self):
        
        
        self.page_width = (self.width + self.leftMargin * 2)
        self.page_height = (self.height + self.bottomMargin * 2)

        styles = getSampleStyleSheet()
        
        styleH = styles['Heading1'] 
        styleH.fontName='simsun'
        
        styleN = styles['Normal']
        styleN.fontName='simsun'        

        # Setting up the frames, frames are use for dynamic content not fixed page elements
        first_page_table_frame = Frame(self.leftMargin, self.bottomMargin, self.width, self.height - 0 * cm, id='small_table')
        
        later_pages_table_frame = Frame(self.leftMargin, self.bottomMargin, self.width, self.height, id='large_table')

        # Creating the page templates
        first_page = PageTemplate(id='FirstPage', frames=[first_page_table_frame], onPage=self.on_first_page)
        later_pages = PageTemplate(id='LaterPages', frames=[later_pages_table_frame], onPage=self.add_default_info)
        self.addPageTemplates([first_page, later_pages])
        # Tell Reportlab to use the other template on the later pages,
        # by the default the first template that was added is used for the first page.        
        story = [NextPageTemplate(['*', 'LaterPages'])]
        story.append(Paragraph('订单编号:{}'.format(self.vals[0][0]), styleH))    
        table_grid = [["物料", "条码号"]]
        # Add the objects
        
        ### Paragraph.wrap = self.wrap
        
    
        for data in self.vals:
            order = data[0]            
            doc = data[1]
            operation = data[2]
            step = data[3]
            whType = data[5]
            partNo = data[6]
            partName = data[7]
            lotNo = data[8]
            if whType == 2:
                sn = data[4].split("|")[0]+"-001"
            else:
                sn = data[4]
            
            
            table_grid.append([partName, sn])

        story.append(Table(table_grid, repeatRows=1, colWidths=[0.6 * self.width, 0.4 * self.width],
                           style=TableStyle([('GRID',(0,0),(-1,-1),0.25,colors.gray), #（起始列，起始行）,(截止列，截止行)
                                             ('ALIGN',(0,0),(-1,0),'CENTER'),
                                             ('BOX', (0,0), (-1,0), 1.0, colors.black),
                                             ('BOX', (0,0), (-1,-1), 1.0, colors.black),                                         
                                             ('FONTNAME', (0,0),(-1,-1),'simsun') 
                                            ])))
        story.append(Spacer(1, 12))
            
        for data in self.vals:
            order = data[0]            
            doc = data[1]
            operation = data[2]
            step = data[3]
            whType = data[5]
            partNo = data[6]
            partName = data[7]
            lotNo = data[8]
            if whType == 2:
                sn = data[4].split("|")[0]+"-001"
            else:
                sn = data[4]
            
            qr_code = qr.QrCodeWidget(sn)
            bounds = qr_code.getBounds()
            
            width = bounds[2] - bounds[0]
            height = bounds[3] - bounds[1]
            H = 120.0
            
            dw = Drawing(H, H, transform=[H/width,0,0,H/height,0,0])
            #dw.add(Paragraph(sn, styleN))
            dw.add(qr_code)
            ### renderPDF.draw(d, c, 15, 405)                
            #story.append(Paragraph("--"*24, styleH)) 
            story.append(Spacer(1, 12))
            ### story.append(Paragraph('工艺规程:{}'.format(doc), styleN))
            ### story.append(Paragraph('工序:{}'.format(operation), styleN))
            ### story.append(Paragraph('工步:{}'.format(step), styleN))
            ### story.append(Paragraph('库存类型:{}'.format(whType), styleN))
            ### story.append(Paragraph('物料号:{}'.format(partNo), styleN))
            ### story.append(Paragraph('物料:{}'.format(partName), styleN))
            ### story.append(Paragraph('批次号:{}'.format(lotNo), styleN))
            
            col1 = ["订单编号","工艺规程","工序","工步","条码号", "库存类型","物料号","物料名称","批次号","数量",""]
            
            table_grid = [["条目","内容","二维码"]]
            
            i = 0
            for val in data:
                if i == 1:
                    table_grid.append([col1[i], val, dw])
                elif i == 0:
                    table_grid.append([col1[i], val, sn])
                elif i== 2 or i == 4:
                    i = i +1
                    continue
                else:
                    table_grid.append([col1[i], val, ""])                    
                i = i +1
            #table_grid.append(["QR","", d])
            story.append(Paragraph('条码号:{}'.format(sn), styleH))
            #story.append(Standard39(sn, ratio=6))
            story.append(Code128(sn, barWidth = inch * 0.012))
            
            story.append(Table(table_grid, repeatRows=1, colWidths=[0.15 * self.width, 0.5 * self.width, 0.35 * self.width],
                               style=TableStyle([('GRID',(0,0),(-1,-1),0.25,colors.gray), #（起始列，起始行）,(截止列，截止行)
                                                 ('ALIGN',(0,0),(-1,0),'CENTER'),
                                                 ('BOX', (0,0), (-1,0), 1.0, colors.black),
                                                 ('BOX', (0,0), (-1,-1), 1.0, colors.black),
                                                 ('SPAN',(-1,2), (-1,-1)),
                                                 ('FONTNAME',(0,0),(-1,-1),'simsun') 
                                                ])))            
            #story.append(d)
        self.build(story)
        
    def wrap(self, availWidth, availHeight):
    # work out widths array for breaking
        self.width = availWidth
        leftIndent = self.style.leftIndent
        first_line_width = availWidth - (leftIndent+self.style.firstLineIndent) - self.style.rightIndent
        later_widths = availWidth - leftIndent - self.style.rightIndent
        try:
            self.blPara = self.breakLinesCJK([first_line_width, later_widths])
        except:
            self.blPara = self.breakLines([first_line_width, later_widths])
        self.height = len(self.blPara.lines) * self.style.leading
        return (self.width, self.height)
    
        

    def on_first_page(self, canvas, doc):
        canvas.saveState()
        # Add the logo and other default stuff
        self.add_default_info(canvas, doc)

        ### canvas.drawString(doc.leftMargin, doc.height, "My address")
        ### canvas.drawString(0.5 * doc.page_width, doc.height, self.their_adress)
        page_num = canvas.getPageNumber()
        text = "Page %s" % page_num
        canvas.drawRightString(200*mm, 20*mm, text)
        canvas.restoreState()

    def add_default_info(self, canvas, doc):
        canvas.saveState()
        canvas.drawCentredString(0.5 * (doc.page_width), doc.page_height - 2.5 * cm, self.filename.split("\\")[-1][:-4])
        page_num = canvas.getPageNumber()
        text = "Page %s" % page_num
        time_str = time.strftime("%Y-%m-%d %H:%M:%S")
        #canvas.drawLiftString(200*mm, 20*mm, "time")
        canvas.drawCentredString(40*mm, 20*mm, time_str)
        canvas.drawRightString(200*mm, 20*mm, text)
        canvas.restoreState()


def main(selected_step):
    
    wb = load_workbook('订单条码生成数据1215.xlsx', data_only=True)

    # sheet 1 to 12 :
    for i in range(2, 3):    

        data = get_data(wb, i)

        for step, vals in data.items():
            
            if selected_step == "0" or selected_step == step.split("-")[-1].split(".")[0]:
                print(step)
                PartListReport(step, vals)
            else:
                pass
                
if __name__ == '__main__':
    
    selected_step = "20"
    
    main(selected_step)