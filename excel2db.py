"""
load table from xlsx

stored procedure:
- sp_load_excel.sql
- sp_load_excel_detail.sql

service: ora_sv

"""
import os
import io
import re

import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf8")

from openpyxl import load_workbook

from nameko.standalone.rpc import ClusterRpcProxy

config = {"AMQP_URI": "pyamqp://guest:guest@10.43.0.240"}


class Loader(object):
    def __init__(self, schema, employee):

        self.merged = []
        self.merged_r_c = {}  # {cell:(row, col)}
        self.ops = [
            "YZP",
            "CJ",
        ]
        self.values = []

        self.schema = schema
        self.employee = employee

        ppattern = r"""([C-Y].*).xlsx"""

        self.compile_obj = re.compile(ppattern)

    def parse_merged_cells(self, sheet, cells_def):

        xcells = cells_def.__dict__
        xsize = cells_def.size

        w = 0
        h = 0
        for row in range(xcells["min_row"], xcells["max_row"] + 1):
            v = sheet.row_dimensions[row].height
            if v == None:
                v = 20
            h = h + v

        for col in range(xcells["min_col"], xcells["max_col"] + 1):
            cname = chr(col + 64)
            v = sheet.column_dimensions[cname].width
            if v == None:
                v = 20
            w = w + v

        i = 0
        for row in range(xcells["min_row"], xcells["max_row"] + 1):
            for col in range(xcells["min_col"], xcells["max_col"] + 1):
                # print("{}-{}".format(chr(col+64), row), end=",")
                if i == 0:
                    self.merged_r_c[(row, col)] = {
                        "columns": xsize["columns"],
                        "rows": xsize["rows"],
                        "width": w,
                        "height": h,
                    }
                else:
                    self.merged.append((row, col))
                i = i + 1

    def get_table_area(self, table_def):
        """
        A26:F32 -> (1[A], 26, 6[F], 32)
        """
        # common variables

        rawstr = r"""([A-Z]+)(\d+):([A-Z]+)(\d+)"""

        compile_obj = re.compile(rawstr)

        try:
            match_obj = compile_obj.search(table_def)

            all_groups = match_obj.groups()

            col_s = ord(match_obj.group(1)) - 64  # col A to int
            row_s = int(match_obj.group(2))
            col_e = ord(match_obj.group(3)) - 64
            row_e = int(match_obj.group(4))
        except:
            print(table_def)
            raise (Exception("error"))
        return col_s, row_s, col_e, row_e

    def process_table(self, sheet, filename, process, characteristic, title, table_def):
        """
        table_def (for example:'A1:D9')
        """

        with ClusterRpcProxy(config) as cluster_rpc:
            
            imgpath = "{}.png".format(characteristic)
            
            if os.path.exists(imgpath):
                pass
            else:
                imgpath = None

            sp_result = cluster_rpc.oracle_service.call_proc(
                ".".join([self.schema, "sp_load_excel"]),
                [process, filename, characteristic, title, imgpath, self.employee, "excel_id:<NUMBER>", "status:<NUMBER>"], "dict"
            )
            
            print(sp_result)
            
            ### excel_id = sp_result[6]            
            ### status = sp_result[7]
            
            excel_id = sp_result["excel_id"]            
            status = sp_result["status"]

            if status == 0 :
                return 0
                
                

            self.values = []
            col_s, row_s, col_e, row_e = self.get_table_area(table_def)
            # print(title)
            pos_row = 0
            pos_col = 0
            for row in range(row_s, row_e + 1):
                pos_row = pos_row + 1
                for col in range(col_s, col_e + 1):
                    pos_col = pos_col + 1
                    if (row, col) in self.merged:
                        # pass the no-value cell in merged cells.
                        pass
                    else:
                        val = sheet.cell(row, col).value
                        if val == None:
                            val = ""
                            editable = 1
                        else:
                            editable = 0
                        if (row, col) in self.merged_r_c:
                            col_span = self.merged_r_c[(row, col)]["columns"]
                            row_span = self.merged_r_c[(row, col)]["rows"]
                            width = self.merged_r_c[(row, col)]["width"]
                            height = self.merged_r_c[(row, col)]["height"]
                        else:
                            col_span = 1
                            row_span = 1
                            cname = chr(col + 64)

                            width = sheet.column_dimensions[cname].width
                            if width == None:
                                width = 20
                            height = sheet.row_dimensions[row].height
                            if height == None:
                                height = 20
                        # print("('{}',{},{},{},{},{},{},{})".format(val, pos_row, pos_col, row_span, col_span, width, height, editable), end="|")
                        ## cell_value, position_row, positon_column, row_span, column_span, cell_width, cell_height, cell_editable, bold?
                        data2db = [
                            excel_id,
                            val,
                            pos_row,
                            pos_col,
                            0,
                            row_span,
                            col_span,
                            width,
                            height,
                            editable,
                            0,
                            self.employee,
                        ]
                        self.values.append(data2db)
                        try:
                            r = cluster_rpc.oracle_service.call_proc(
                                ".".join([self.schema, "sp_load_excel_detail"]), data2db
                            )
                            ### print(data2db)
                            print(r)
                            print(".", end="")
                        except:
                            print("{},{},{}".format(fn, characteristic, title))
                            raise (Exception("error"))
                        # print(val, end=" | ")
                # print()
            # print(self.values)

    def load(self, filename, sheet_name):

        wb = load_workbook(filename=filename)

        sheet = wb[sheet_name]

        max_row = sheet.max_row

        ## get merged cells
        mcells = sheet.merged_cells

        ## 处理合并的cells
        for item in mcells:
            self.parse_merged_cells(sheet, item)

        ## 逐行遍历
        for i in range(1, max_row + 1):
            # print(i)
            val = str(sheet.cell(i, 1).value)
            tt = val.split("-")
            if tt[0] in self.ops and len(tt) > 1:  ##遇到 55.2 这种
                # print(val)
                ## 下一行为表格Title
                title = sheet.cell(i + 1, 1).value

                ## 获得质检项table区域，如A3:D10
                table_def = sheet.cell(i, 2).value

                ## 处理定义的Table区域内容
                match_obj = self.compile_obj.search(filename)
                process = match_obj.group(1)
                characteristic = val
                self.process_table(sheet, filename, process, characteristic, title, table_def)


if __name__ == "__main__":

    employee = "mabo"

    schema = "flxuser"

    loader = Loader(schema, employee)

    fns = [
        ### "1-本体YZP-C1H-E1-2018-0005.xlsx",
        ### "2-本体YZP-C1H-E1-2018-0013.xlsx",
        ### "4-部件CJ-1000AXH.E1QT.462.xlsx",
        "2.YZP-C1H-E1-2018-0013.xlsx",
        ### "5-部件YZP-C1H-E1-2017-0004.xlsx",
        ### "6-部件YZP-C1H-E1-2018-0007.xlsx",
        ### "7-部件CJ-1000AXH.E1QT.447.xlsx",
        ### "9-总装YZP-C1H-E1-2017-0003.xlsx",
        ### "10-总装YZP-C1H-E1-2017-0001.xlsx",
        ### "11★-总装YZP-C1H-E1-2017-0002.xlsx",
        ### "12-平衡YZP-C1H-E1-2018-0012.xlsx",
        ### "13-平衡YZP-C1H-E1-2018-0011.xlsx",
        ### "14-平衡CJ-1000AXH.E1QT.460.xlsx",
        ### "15-平衡CJ-1000AXH.E1QT.458.xlsx",
    ]
    for fn in fns:
        loader.load(fn, "Sheet2")
